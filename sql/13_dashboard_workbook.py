import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

DB = "collections_analysis.db"
OUT = "../dashboards/collections_dashboard.xlsx"

# ---- palette (matches HTML dashboard) ----
NAVY = "1E2A3A"
TEAL = "0E6E63"
TEAL_SOFT = "DCEAE7"
AMBER = "B9812E"
AMBER_SOFT = "F3E6CC"
RED = "A8402F"
RED_SOFT = "F3DED8"
PAPER = "F2F1EC"
INK = "1C2126"
INK_SOFT = "5B6470"
LINE = "D9D6CC"
WHITE = "FFFFFF"

FONT_NAME = "Arial"

title_font = Font(name=FONT_NAME, size=18, bold=True, color=WHITE)
eyebrow_font = Font(name=FONT_NAME, size=9, bold=True, color="9FB3AE")
sub_font = Font(name=FONT_NAME, size=10, italic=True, color="B7BEC9")
header_font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
label_font = Font(name=FONT_NAME, size=9, bold=True, color=INK_SOFT)
value_font = Font(name=FONT_NAME, size=16, bold=True, color=INK)
body_font = Font(name=FONT_NAME, size=10, color=INK)
note_font = Font(name=FONT_NAME, size=9, italic=True, color=INK_SOFT)
section_font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)

navy_fill = PatternFill("solid", fgColor=NAVY)
teal_fill = PatternFill("solid", fgColor=TEAL)
paper_fill = PatternFill("solid", fgColor=PAPER)
white_fill = PatternFill("solid", fgColor=WHITE)
amber_soft_fill = PatternFill("solid", fgColor=AMBER_SOFT)
red_soft_fill = PatternFill("solid", fgColor=RED_SOFT)
teal_soft_fill = PatternFill("solid", fgColor=TEAL_SOFT)

thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, ncols, fill=navy_fill, font=header_font):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_title_block(ws, title, subtitle, span=6, rows=3):
    ws.merge_cells(start_row=1, start_column=1, end_row=rows, end_column=span)
    cell = ws.cell(row=1, column=1)
    cell.value = f"{title}\n{subtitle}"
    cell.font = Font(name=FONT_NAME, size=13, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    cell.fill = navy_fill
    for r in range(1, rows + 1):
        for c in range(1, span + 1):
            ws.cell(row=r, column=c).fill = navy_fill
    ws.row_dimensions[1].height = 22
    if rows > 1:
        ws.row_dimensions[2].height = 16

con = sqlite3.connect(DB)

# ================= Pull real numbers from the golden layer =================
monthly = pd.read_sql_query("""
    SELECT month,
           ROUND(SUM(successful_payment_amount), 2) AS successful_recovery,
           SUM(successful_payment_count) AS successful_payments,
           SUM(payment_count) AS total_payments,
           ROUND(100.0 * SUM(successful_payment_count) / NULLIF(SUM(payment_count), 0), 2) AS payment_success_rate_pct,
           SUM(answered_calls) AS answered_calls,
           SUM(call_attempts) AS call_attempts,
           SUM(ptp_count) AS ptp_count,
           SUM(kept_ptp_count) AS kept_ptp_count
    FROM account_month
    GROUP BY month ORDER BY month
""", con)
monthly["recovery_mom_pct"] = monthly["successful_recovery"].pct_change() * 100

risk = pd.read_sql_query("""
    SELECT risk_segment, COUNT(DISTINCT account_id) AS accounts,
           ROUND(SUM(successful_payment_amount), 2) AS successful_recovery,
           ROUND(SUM(successful_payment_amount) / COUNT(DISTINCT account_id), 2) AS recovery_per_account
    FROM account_month GROUP BY risk_segment ORDER BY successful_recovery DESC
""", con)

dpd = pd.read_sql_query("""
    SELECT CASE WHEN dpd = 0 THEN '0' WHEN dpd BETWEEN 1 AND 30 THEN '1-30'
                WHEN dpd BETWEEN 31 AND 60 THEN '31-60' WHEN dpd BETWEEN 61 AND 90 THEN '61-90'
                ELSE '90+' END AS dpd_bucket,
           COUNT(DISTINCT account_id) AS accounts,
           ROUND(SUM(successful_payment_amount), 2) AS successful_recovery
    FROM account_month GROUP BY dpd_bucket
    ORDER BY CASE dpd_bucket WHEN '0' THEN 1 WHEN '1-30' THEN 2 WHEN '31-60' THEN 3 WHEN '61-90' THEN 4 ELSE 5 END
""", con)

channel = pd.read_sql_query("""
    SELECT recommended_channel AS channel, COUNT(*) AS targets,
           SUM(CASE WHEN status='CONTACTED' THEN 1 ELSE 0 END) AS contacted,
           ROUND(100.0 * SUM(CASE WHEN status='CONTACTED' THEN 1 ELSE 0 END) / COUNT(*), 2) AS contact_rate_pct
    FROM daily_targeting GROUP BY recommended_channel ORDER BY contact_rate_pct DESC
""", con)

total_accounts = pd.read_sql_query("SELECT COUNT(*) n FROM accounts", con).iloc[0]["n"]
total_outstanding = pd.read_sql_query("SELECT SUM(outstanding_amount) t FROM accounts", con).iloc[0]["t"]
raw_payments = pd.read_sql_query("SELECT COUNT(*) n FROM payments", con).iloc[0]["n"]

con.close()

full_months = monthly[monthly["month"] != "2026-08"].copy()
jan = full_months.iloc[0]["successful_recovery"]
jul = full_months.iloc[-1]["successful_recovery"]
overall_change = (jul - jan) / jan * 100
mom_vals = full_months["recovery_mom_pct"].dropna()
avg_mom = mom_vals.mean()
feb_mar = full_months.iloc[2]["recovery_mom_pct"]

wb = Workbook()

# ============================================================= COVER =============================================================
cov = wb.active
cov.title = "Cover"
cov.sheet_view.showGridLines = False
cov.sheet_properties.tabColor = NAVY
autosize(cov, [3, 34, 34, 20, 20, 20])

cov.merge_cells("A1:F1")
cov["A1"] = "COLLECTIONS ANALYTICS · JAN\u2013AUG 2026 · GOLDEN DATASET v1"
cov["A1"].font = Font(name=FONT_NAME, size=9, bold=True, color=TEAL)
cov.row_dimensions[1].height = 18

cov.merge_cells("A2:F2")
cov["A2"] = "Did recovery actually improve by 11%?"
cov["A2"].font = Font(name=FONT_NAME, size=20, bold=True, color=NAVY)
cov.row_dimensions[2].height = 30

cov.merge_cells("A3:F3")
cov["A3"] = "Independent audit of the reported month-on-month recovery claim, its drivers, data quality, and the \u20b910 Cr allocation decision."
cov["A3"].font = Font(name=FONT_NAME, size=10, italic=True, color=INK_SOFT)
cov.row_dimensions[3].height = 20

cov.merge_cells("A5:C6")
cov["A5"] = "VERDICT ON THE 11% CLAIM"
cov["A5"].font = Font(name=FONT_NAME, size=9, bold=True, color=WHITE)
cov["A5"].fill = PatternFill("solid", fgColor=RED)
cov["A5"].alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
for r in range(5, 7):
    for c in range(1, 4):
        cov.cell(row=r, column=c).fill = PatternFill("solid", fgColor=RED)
cov.merge_cells("D5:F6")
cov["D5"] = f"NOT SUPPORTED \u2014 cherry-picked month\n7-month trend is flat ({overall_change:+.2f}%), avg MoM {avg_mom:+.2f}%"
cov["D5"].font = Font(name=FONT_NAME, size=11, bold=True, color=RED)
cov["D5"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

kpi_labels = ["Total accounts", "Successful payments", "Successful recovery (7mo)", "Recovery rate", "Kept PTPs", "PTP\u2192payment conversion"]
kpi_values = [
    f"{total_accounts:,}",
    f"{int(full_months['successful_payments'].sum()):,}",
    f"\u20b9{full_months['successful_recovery'].sum()/1e7:,.2f} Cr",
    f"{full_months['successful_recovery'].sum()/total_outstanding*100:.2f}%",
    f"{int(full_months['kept_ptp_count'].sum()):,}",
    "20.67%",
]
start_row = 8
cov.cell(row=start_row, column=1, value="KEY NUMBERS").font = section_font
for i, (lab, val) in enumerate(zip(kpi_labels, kpi_values)):
    r = start_row + 1 + i
    cov.cell(row=r, column=1, value=lab).font = label_font
    cov.cell(row=r, column=1).alignment = Alignment(indent=1)
    cell = cov.cell(row=r, column=2, value=val)
    cell.font = Font(name=FONT_NAME, size=12, bold=True, color=TEAL)
    for c in range(1, 3):
        cov.cell(row=r, column=c).border = border
        cov.cell(row=r, column=c).fill = white_fill

nav_row = start_row + len(kpi_labels) + 3
cov.cell(row=nav_row, column=1, value="CONTENTS").font = section_font
sheets_desc = [
    ("Monthly Recovery", "Golden-layer monthly KPIs with MoM % change"),
    ("Claim Audit", "Reported vs. independently recalculated recovery trend"),
    ("Risk & DPD", "Recovery cut by risk segment and delinquency bucket"),
    ("Channel Performance", "Contact rate and recovery by recommended channel"),
    ("Data Quality", "Findings from raw \u2192 golden reconciliation"),
    ("Investment Options", "Evidence available for each \u20b910 Cr allocation option"),
]
for i, (name, desc) in enumerate(sheets_desc):
    r = nav_row + 1 + i
    cov.cell(row=r, column=1, value=name).font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
    cov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    cov.cell(row=r, column=2, value=desc).font = body_font

foot_row = nav_row + len(sheets_desc) + 3
cov.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=6)
cov.cell(row=foot_row, column=1,
         value="Source: sql/collections_analysis.db (account_month golden table). Full narrative in reports/executive_memo.md and dashboards/collections_recovery_dashboard.html. August 2026 excluded from trend statistics (partial month, 8 of 31 days).").font = note_font

# ============================================================= MONTHLY RECOVERY =============================================================
ws = wb.create_sheet("Monthly Recovery")
ws.sheet_view.showGridLines = False
write_title_block(ws, "Monthly Recovery", "Golden-layer KPIs by month \u2014 account_month table", span=7, rows=2)
headers = ["Month", "Successful Recovery (\u20b9)", "Successful Payments", "Total Payments", "Payment Success Rate", "Answered Calls", "MoM Change"]
hr = 4
for i, h in enumerate(headers, start=1):
    ws.cell(row=hr, column=i, value=h)
style_header_row(ws, hr, len(headers), fill=teal_fill)
ws.row_dimensions[hr].height = 26

for i, row in monthly.iterrows():
    r = hr + 1 + i
    ws.cell(row=r, column=1, value=row["month"]).font = body_font
    c2 = ws.cell(row=r, column=2, value=row["successful_recovery"])
    c2.number_format = '"\u20b9"#,##0'
    ws.cell(row=r, column=3, value=int(row["successful_payments"]))
    ws.cell(row=r, column=4, value=int(row["total_payments"]))
    c5 = ws.cell(row=r, column=5, value=row["payment_success_rate_pct"] / 100)
    c5.number_format = "0.0%"
    ws.cell(row=r, column=6, value=int(row["answered_calls"]))
    mom = row["recovery_mom_pct"]
    c7 = ws.cell(row=r, column=7, value=(mom / 100) if pd.notna(mom) else None)
    c7.number_format = "+0.0%;-0.0%"
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        cell.fill = paper_fill if i % 2 else white_fill
        if c != 1 and cell.font is None:
            cell.font = body_font
    if row["month"] == "2026-08":
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = amber_soft_fill
        ws.cell(row=r, column=1, value="2026-08*").font = Font(name=FONT_NAME, italic=True, size=10)

# conditional format on MoM column
last_data_row = hr + len(monthly)
mom_range = f"G{hr+1}:G{last_data_row}"
ws.conditional_formatting.add(mom_range, CellIsRule(operator="greaterThan", formula=["0"], fill=teal_soft_fill))
ws.conditional_formatting.add(mom_range, CellIsRule(operator="lessThan", formula=["0"], fill=red_soft_fill))

note_row = last_data_row + 2
ws.cell(row=note_row, column=1, value="* August 2026 is a partial month (8 of 31 days) \u2014 excluded from all trend and verdict statistics.").font = note_font
ws.cell(row=note_row + 1, column=1, value="Recovery rate denominator held constant at total portfolio outstanding (\u20b91,048.9 Cr) across months so the ratio isolates recovery changes, not portfolio size.").font = note_font
autosize(ws, [12, 22, 18, 16, 18, 16, 14])

# ============================================================= CLAIM AUDIT =============================================================
ws2 = wb.create_sheet("Claim Audit")
ws2.sheet_view.showGridLines = False
write_title_block(ws2, "Auditing the \u201811% recovery improvement\u2019 claim", "Reported vs. independently recalculated, Jan\u2013Jul 2026", span=4, rows=2)

rows2 = [
    ("Reported by business", "+11.0%", "Undisclosed method; matches Feb\u2192Mar month-pair only", AMBER_SOFT),
    ("Closest real match", f"{feb_mar:+.2f}%", "Feb \u2192 Mar month-pair \u2014 the one positive outlier out of six pairs", None),
    ("Independently recalculated (avg. of all pairs)", f"{avg_mom:+.2f}%", "Mean MoM % change, successful payment amount, Jan\u2013Jul", TEAL_SOFT),
    ("Jan vs. Jul, net change", f"{overall_change:+.2f}%", f"\u20b9{jan/1e7:.2f} Cr \u2192 \u20b9{jul/1e7:.2f} Cr \u2014 effectively unchanged", TEAL_SOFT),
    ("Verdict", "NOT SUPPORTED", "The 11% figure is a single favorable month-pair, not a sustained trend", RED_SOFT),
]
headers2 = ["Basis", "MoM / Net Change", "Method / Note"]
hr2 = 4
for i, h in enumerate(headers2, start=1):
    ws2.cell(row=hr2, column=i, value=h)
style_header_row(ws2, hr2, 3, fill=navy_fill)
ws2.row_dimensions[hr2].height = 22
for i, (label, val, note, fill) in enumerate(rows2):
    r = hr2 + 1 + i
    ws2.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, size=10, bold=True, color=INK)
    vcell = ws2.cell(row=r, column=2, value=val)
    vcell.font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
    vcell.alignment = Alignment(horizontal="center")
    ws2.cell(row=r, column=3, value=note).font = body_font
    for c in range(1, 4):
        cell = ws2.cell(row=r, column=c)
        cell.border = border
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
autosize(ws2, [34, 18, 60])
ws2.row_dimensions[hr2 + len(rows2) + 2].height = 10
nr2 = hr2 + len(rows2) + 3
ws2.merge_cells(start_row=nr2, start_column=1, end_row=nr2, end_column=3)
ws2.cell(row=nr2, column=1,
         value="Regressing monthly recovery against month index gives R\u00b2 = 0.004 \u2014 month explains essentially none of the variance. Full decomposition in reports/executive_memo.md.").font = note_font

# ============================================================= RISK & DPD =============================================================
ws3 = wb.create_sheet("Risk & DPD")
ws3.sheet_view.showGridLines = False
write_title_block(ws3, "Recovery by Risk Segment and DPD Bucket", "Golden layer \u2014 account_month table", span=4, rows=2)

ws3.cell(row=4, column=1, value="RISK SEGMENT").font = section_font
headers3 = ["Risk Segment", "Accounts", "Successful Recovery (\u20b9)", "Recovery / Account (\u20b9)"]
hr3 = 5
for i, h in enumerate(headers3, start=1):
    ws3.cell(row=hr3, column=i, value=h)
style_header_row(ws3, hr3, 4, fill=teal_fill)
for i, row in risk.iterrows():
    r = hr3 + 1 + i
    ws3.cell(row=r, column=1, value=row["risk_segment"]).font = body_font
    ws3.cell(row=r, column=2, value=int(row["accounts"]))
    c = ws3.cell(row=r, column=3, value=row["successful_recovery"]); c.number_format = '"\u20b9"#,##0'
    c = ws3.cell(row=r, column=4, value=row["recovery_per_account"]); c.number_format = '"\u20b9"#,##0'
    for cc in range(1, 5):
        ws3.cell(row=r, column=cc).border = border
        ws3.cell(row=r, column=cc).fill = paper_fill if i % 2 else white_fill

dpd_start = hr3 + len(risk) + 3
ws3.cell(row=dpd_start, column=1, value="DPD BUCKET").font = section_font
headers3b = ["DPD Bucket", "Accounts", "Successful Recovery (\u20b9)"]
hr3b = dpd_start + 1
for i, h in enumerate(headers3b, start=1):
    ws3.cell(row=hr3b, column=i, value=h)
style_header_row(ws3, hr3b, 3, fill=teal_fill)
for i, row in dpd.iterrows():
    r = hr3b + 1 + i
    ws3.cell(row=r, column=1, value=row["dpd_bucket"]).font = body_font
    ws3.cell(row=r, column=2, value=int(row["accounts"]))
    c = ws3.cell(row=r, column=3, value=row["successful_recovery"]); c.number_format = '"\u20b9"#,##0'
    for cc in range(1, 4):
        ws3.cell(row=r, column=cc).border = border
        ws3.cell(row=r, column=cc).fill = paper_fill if i % 2 else white_fill
autosize(ws3, [18, 14, 22, 20])

# ============================================================= CHANNEL PERFORMANCE =============================================================
ws4 = wb.create_sheet("Channel Performance")
ws4.sheet_view.showGridLines = False
write_title_block(ws4, "Targeting Contact Rate by Channel", "daily_targeting table \u2014 all campaigns, full period", span=4, rows=2)
headers4 = ["Channel", "Targets", "Contacted", "Contact Rate"]
hr4 = 4
for i, h in enumerate(headers4, start=1):
    ws4.cell(row=hr4, column=i, value=h)
style_header_row(ws4, hr4, 4, fill=teal_fill)
for i, row in channel.iterrows():
    r = hr4 + 1 + i
    ws4.cell(row=r, column=1, value=row["channel"]).font = body_font
    ws4.cell(row=r, column=2, value=int(row["targets"]))
    ws4.cell(row=r, column=3, value=int(row["contacted"]))
    c = ws4.cell(row=r, column=4, value=row["contact_rate_pct"] / 100); c.number_format = "0.00%"
    for cc in range(1, 5):
        ws4.cell(row=r, column=cc).border = border
        ws4.cell(row=r, column=cc).fill = paper_fill if i % 2 else white_fill
nr4 = hr4 + len(channel) + 2
ws4.merge_cells(start_row=nr4, start_column=1, end_row=nr4, end_column=4)
ws4.cell(row=nr4, column=1, value="Contact-rate differences across channels are within ~1 point of each other \u2014 not a sufficient basis alone for a large single-channel investment. See Driver tab in the HTML dashboard for recovery-per-target by channel, which uses a 30-day attribution window.").font = note_font
autosize(ws4, [16, 14, 14, 14])

# ============================================================= DATA QUALITY =============================================================
ws5 = wb.create_sheet("Data Quality")
ws5.sheet_view.showGridLines = False
write_title_block(ws5, "Data Quality Findings", "Raw \u2192 flagged/corrected \u2192 golden reconciliation", span=5, rows=2)
headers5 = ["Table", "Total Rows", "Flagged / Affected", "% Affected", "Treatment"]
hr5 = 4
for i, h in enumerate(headers5, start=1):
    ws5.cell(row=hr5, column=i, value=h)
style_header_row(ws5, hr5, 5, fill=navy_fill)
dq_rows = [
    ("Payments", 25500, "500 duplicate payment_ids (486 exact-duplicate rows, \u20b92.59 Cr double-counted SUCCESS value)", 2.0, "Deduplicated on payment_id, first occurrence kept"),
    ("Accounts", 30000, "13,029 accounts where outstanding_amount > principal_amount", 43.4, "Flagged, not corrected \u2014 no business rule available"),
    ("Borrowers", 30000, "27,503 rows with conflicting name/phone/email/city/state under one borrower_id", 91.7, "borrower_id treated as unreliable identity key; flagged"),
    ("Agents (master)", 30000, "1,000 of 1,000 agent_ids show conflicting employee_code/vendor_id/team/status across rows", 100.0, "Requires effective-dated join \u2014 not yet implemented"),
    ("Calls", 90079, "1,827 missing agent_id; 158 conflicting duplicate call_ids", 2.2, "Retained, flagged as incomplete"),
    ("Call attempts", 120000, "2,400 missing vendor_id", 2.0, "Retained, flagged as incomplete"),
    ("Account status history", 60000, "30,191 records with recorded_at earlier than event_at", 50.3, "Retained, flagged as backdated/late-arriving"),
]
for i, (table, total, affected, pct, treatment) in enumerate(dq_rows):
    r = hr5 + 1 + i
    ws5.cell(row=r, column=1, value=table).font = Font(name=FONT_NAME, size=10, bold=True)
    ws5.cell(row=r, column=2, value=total).number_format = "#,##0"
    ws5.cell(row=r, column=3, value=affected).font = body_font
    ws5.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="center")
    c = ws5.cell(row=r, column=4, value=pct / 100); c.number_format = "0.0%"
    ws5.cell(row=r, column=5, value=treatment).font = body_font
    ws5.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="center")
    for cc in range(1, 6):
        ws5.cell(row=r, column=cc).border = border
        ws5.cell(row=r, column=cc).fill = paper_fill if i % 2 else white_fill
    ws5.row_dimensions[r].height = 34
autosize(ws5, [18, 12, 46, 12, 40])

# ============================================================= INVESTMENT OPTIONS =============================================================
ws6 = wb.create_sheet("Investment Options")
ws6.sheet_view.showGridLines = False
write_title_block(ws6, "Where should the \u20b910 Cr go?", "Evidence available per option \u2014 none has cost data", span=4, rows=2)
headers6 = ["Option", "What the data shows", "Cost data available?", "Confidence"]
hr6 = 4
for i, h in enumerate(headers6, start=1):
    ws6.cell(row=hr6, column=i, value=h)
style_header_row(ws6, hr6, 4, fill=navy_fill)
inv_rows = [
    ("Better telephony infrastructure", "Contact rate ~19\u201320% flat across all channels/vendors \u2014 no vendor stands out as a bottleneck.", "No", "Hypothesis"),
    ("More collection agents", "Recovery/agent-hour flat at ~\u20b916.1k\u201317.0k month to month; no evidence agents are volume-constrained.", "No", "Hypothesis"),
    ("AI voice automation", "No current agentic-voice channel exists in the data to benchmark against.", "No", "Hypothesis"),
    ("Better borrower targeting", "Attribution-window sensitivity (1,338\u20139,505 matched payments) suggests current logic is under-specified.", "No", "Correlation"),
    ("WhatsApp / digital engagement", "Highest call volume (23,422) but average, not above-average, contact/recovery outcomes.", "No", "Correlation"),
    ("Field operations", "Lowest contact volume but comparable per-target recovery (\u20b975.3k) \u2014 plausibly most expensive per contact.", "No", "Hypothesis"),
]
conf_fill = {"Fact": TEAL_SOFT, "Strong evidence": TEAL_SOFT, "Correlation": AMBER_SOFT, "Hypothesis": RED_SOFT}
for i, (opt, note, cost, conf) in enumerate(inv_rows):
    r = hr6 + 1 + i
    ws6.cell(row=r, column=1, value=opt).font = Font(name=FONT_NAME, size=10, bold=True)
    ws6.cell(row=r, column=2, value=note).font = body_font
    ws6.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    ws6.cell(row=r, column=3, value=cost).alignment = Alignment(horizontal="center")
    cc4 = ws6.cell(row=r, column=4, value=conf)
    cc4.alignment = Alignment(horizontal="center")
    cc4.fill = PatternFill("solid", fgColor=conf_fill.get(conf, WHITE))
    for cc in range(1, 4):
        ws6.cell(row=r, column=cc).border = border
        ws6.cell(row=r, column=cc).fill = paper_fill if i % 2 else white_fill
    ws6.cell(row=r, column=4).border = border
    ws6.row_dimensions[r].height = 40
autosize(ws6, [28, 60, 16, 14])
nr6 = hr6 + len(inv_rows) + 2
ws6.merge_cells(start_row=nr6, start_column=1, end_row=nr6, end_column=4)
ws6.cell(row=nr6, column=1,
         value="No cost, salary, campaign-spend, or vendor-pricing field exists in any of the 17 source tables \u2014 ROI and break-even are not computable for any option on this data. Recommendation: fund a controlled 2\u20133 month pilot with cost tracked from day one, sized well below \u20b910 Cr. Full reasoning in reports/investment_recommendation.md.").font = note_font

wb.save(OUT)
print("Saved", OUT)
